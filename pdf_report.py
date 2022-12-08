import csv
from _datetime import datetime
import re
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.styles import Font
from openpyxl.styles.numbers import BUILTIN_FORMATS
from jinja2 import Environment, FileSystemLoader
import pdfkit
"""
Файл, реализующий показ статистики
"""

currencyToRub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055}


class Vacancy:
    """Класс для представления вакансии

    Attribute:
        name (str): Название вакансии
        salary (str): Зарплата
        area_name (str): Местоположение
        published_at (str): Дата публикации
    """
    def __init__(self, name, salary, area_name, published_at):
        """Инициализирует объект Vacancy

        Args:
        :param name (str): Название вакансии
        :param salary (str): Зарплата
        :param area_name (str): Местоположение
        :param published_at (str): Дата публикации
        """
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at


class Salary:
    """Класс для представления зарплаты

    Attribute:
        salary_from(int): Нижняя граница оклада
        salary_to(int): Верхняя граница оклада
        salary_currency(str): Валюта
    """
    def __init__(self, salary_from, salary_to, salary_currency):
        """Инициализирует объект Salary

        Args:
        :param salary_from(int): Нижняя граница оклада
        :param salary_to(int): Верхняя граница оклада
        :param salary_currency(str): Валюта

        >>> type(Salary(11.0, 21.4, 'RUR')).__name__
        'Salary'
        >>> Salary(11.0, 21.4, 'RUR').salary_from
        11.0
        >>> Salary(11.0, 21.4, 'RUR').salary_to
        21.4
        >>> Salary(11.0, 21.4, 'RUR').salary_currency
        'RUR'
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        self.salary_ru = int((float(self.salary_from) + float(self.salary_to)) / 2) * currencyToRub[
            self.salary_currency]

    def get_salary_ru(self):
        """Вычисляет ср зарплату и переводит ее в рубли

        :return: float salary_ru

        >>> Salary(11.0, 21.4, 'RUR').get_salary_ru()
        16
        >>> Salary(10, 20, 'RUR').get_salary_ru()
        15
        >>> Salary(10, 30.0, 'RUR').get_salary_ru()
        20
        >>> Salary(10, 30.0, 'EUR').get_salary_ru()
        1198.0
        """
        return self.salary_ru


class DataSet:
    """Класс для составления списка вакансий

    Attribute:
        file_name(str): Имя файла
    """
    def __init__(self, file_name):
        """Инициализирует объект DataSet

        :param file_name: (str): Имя файла
        """
        self.file_name = file_name
        self.vacancies_objects = DataSet.prepare(file_name)

    @staticmethod
    def csv_reader(filename):
        """Парсер csv файла. Если файл не пустой, то он делится на названия колонок и сами записи вакансий

        :param filename: (str): Имя файла
        :return:
            clmns (list[str]): названия колонок,
            lines (list[list[str]]): записи вакансий
        """
        with open(filename, encoding="utf-8-sig") as f:
            data = [x for x in csv.reader(f)]
        try:
            clmns = data[0]
            lines = data[1:]
            return clmns, lines
        except FileNotFoundError:
            print("Пустой файл")
            exit()

    @staticmethod
    def prepare(filename):
        """Фильтрует вакансии и создает список вакансий(словарей), где key - название колонки, а value - значение

        :param filename: (str): Имя файла
        :return: vac(list[Vacancy]): список вакансий
        """
        clmns, lines = DataSet.csv_reader(filename)
        filtred = [i for i in lines if len(i) == len(clmns) and '' not in i]
        vac = []
        for line in filtred:
            dct = {}
            for x in range(0, len(line)):
                if line[x].count('\n') > 0:
                    read = [DataSet.remove_tags(el) for el in line[x].split('\n')]
                else:
                    read = DataSet.remove_tags(line[x])
                dct[clmns[x]] = read

            vac.append(Vacancy(dct['name'], Salary(dct['salary_from'],
                                                   dct['salary_to'], dct['salary_currency']), dct['area_name'],
                               dct['published_at']))
        return vac

    @staticmethod
    def remove_tags(args):
        """Удаляет html теги

        :param args: (list[str]) вакансия
        :return: (list[str]) отфильтровонная вакансия
        """
        return " ".join(re.sub(r"\<[^>]*\>", "", args).split())


class InputConnect:
    """Класс для получения входных данных и вывода статистики

    Attribute:

    """
    def __init__(self):
        """Инициализирует объект InputConnect

        Вызывает метод prepare класса DataSet для получения списка вакансий
        Вызывает метод print класса InputConnect для вывода статистики
        """
        params = InputConnect.get_prms()
        data = DataSet.prepare(params[0])
        InputConnect.print(data, params[1])

    @staticmethod
    def get_prms():
        """Реализует получение входных данных

        :return:
            file_name(str): имя файла
            vacancy(str): название профессии
        """
        file_name = input("Введите название файла: ")
        vacancy = input("Введите название профессии: ")
        return file_name, vacancy

    @staticmethod
    def first_el(dic):
        """Выбирает только первые 10 элементов

        :param dic: (dict): словарь, содержащий статистику
        :return: new_dic (dict): новый словарь статистики(первые 10 пар dic)
        """
        new_dic = {}
        i = 0
        for key, value in dic.items():
            new_dic[key] = value
            i += 1
            if i == 10:
                break
        return new_dic

    @staticmethod
    def print(dic_vacancies, vac_name):
        """Выводит статистику и вызывает Класс Report

        :param dic_vacancies: (list[Vacancy]) список вакансий
        :param vac_name: (str) название профессии
        """
        years = set()
        for vacancy in dic_vacancies:
            years.add(int(vacancy.published_at[:4]))
        years = sorted(list(years))
        years = list(range(min(years), max(years) + 1))

        salary_years = {year: [] for year in years}
        vacs_years = {year: 0 for year in years}
        vac_salary_years = {year: [] for year in years}
        vac_count_years = {year: 0 for year in years}

        for vacancy in dic_vacancies:
            year = int(vacancy.published_at[:4])
            salary_years[year].append(vacancy.salary.get_salary_ru())
            vacs_years[year] += 1
            if vac_name in vacancy.name:
                vac_salary_years[year].append(vacancy.salary.get_salary_ru())
                vac_count_years[year] += 1

        salary_years = {key: int(sum(value) / len(value)) if len(value) != 0 else 0
                        for key, value in salary_years.items()}
        vac_salary_years = {key: int(sum(value) / len(value)) if len(value) != 0 else 0
                            for key, value in vac_salary_years.items()}

        area_dic = {}
        for vacancy in dic_vacancies:
            if vacancy.area_name in area_dic:
                area_dic[vacancy.area_name].append(vacancy.salary.get_salary_ru())
            else:
                area_dic[vacancy.area_name] = [vacancy.salary.get_salary_ru()]

        area_salary = [x for x in area_dic.items() if len(x[1]) / len(dic_vacancies) > 0.01]
        sort_area_salary = sorted(area_salary, key=lambda item: sum(item[1]) / len(item[1]), reverse=True)
        res_sort_area_salary = {item[0]: int(sum(item[1]) / len(item[1])) for item in sort_area_salary}

        fract_vac_area = {
            key: round(len(value) / len(dic_vacancies), 4) if len(value) / len(dic_vacancies) > 0.01 else 0
            for key, value in area_dic.items()}
        fract_vac_area = {key: value for key, value in fract_vac_area.items() if value != 0}
        sort_fract_vac_area = sorted(fract_vac_area.items(), key=lambda item: item[1], reverse=True)
        res_sort_fract_vac_area = {k: v for k, v in sort_fract_vac_area}

        print('Динамика уровня зарплат по годам: {}'.format(salary_years))
        print('Динамика количества вакансий по годам: {}'.format(vacs_years))
        print('Динамика уровня зарплат по годам для выбранной профессии: {}'.format(vac_salary_years))
        print('Динамика количества вакансий по годам для выбранной профессии: {}'.format(vac_count_years))
        print('Уровень зарплат по городам (в порядке убывания): {}'.format(InputConnect.first_el(res_sort_area_salary)))
        print(
            'Доля вакансий по городам (в порядке убывания): {}'.format(InputConnect.first_el(res_sort_fract_vac_area)))

        res = [salary_years, vac_salary_years, vacs_years, vac_count_years, InputConnect.first_el(
            res_sort_area_salary), InputConnect.first_el(res_sort_fract_vac_area)]

        Report(res, vac_name)


class Report:
    """Класс для реализации всей статистики

    Attribute:
        info(list[dict]): список статистики
        vac_name(str): название профессии
    """
    def __init__(self, info, vac_name):
        """Инициализирует объект Report и вызывает методы класса Report для создания статистических файлов

        :param info: (list[dict]): список статистики
        :param vac_name: (str): название профессии
        """
        Report.vac_name = vac_name
        Report.info = info
        Report.report_excel(info, vac_name)
        Report.generate_image(info, vac_name)
        Report.generate_pdf(info, vac_name)

    @staticmethod
    def sym_n(dct):
        """Проходится по словарю и заменяет '-' или ' ' на '\n' в ключах словаря

        :param dct: (dict) словарь статистики
        :return: dct: (dict) новый словарь статистики
        """
        for key in list(dct.keys()):
            if '-' in key:
                new = key.replace('-', '\n')
                dct[new] = dct[key]
                del dct[key]
            if ' ' in key:
                new = key.replace(' ', '\n')
                dct[new] = dct[key]
                del dct[key]

        dct = {k: v for k, v in sorted(dct.items(), key=lambda item: item[1], reverse=True)}
        return dct

    @staticmethod
    def plus_other(dct):
        """Добавляет в словарь ключ 'Другие' со значением остатка долей

        :param dct: (dict) словарь статистики
        :return: dct: (dict) новый словарь статистики
        """
        dct['Другие'] = 1 - sum(dct.values())
        dct = {k: v for k, v in sorted(dct.items(), key=lambda item: item[1], reverse=True)}
        return dct

    @staticmethod
    def generate_image(info, vac):
        """Генерирует файл png с графиками по статистике

        :param info: (list[dict]): список статистики
        :param vac: (str): название профессии
        """
        salary_years = info[0]
        vac_salary_years = info[1]
        vacs_years = info[2]
        vac_count_years = info[3]
        res_sort_area_salary = info[4]
        res_sort_fract_vac_area = info[5]

        width = 0.3
        x_nums = np.arange(len(salary_years.keys()))
        x_list1 = x_nums - width / 2
        x_list2 = x_nums + width / 2

        fig = plt.figure()

        # 1 - Уровень зарплат по годам
        ax = fig.add_subplot(221)
        ax.set_title('Уровень зарплат по годам')
        ax.bar(x_list1, salary_years.values(), width, label='средняя з/п')
        ax.bar(x_list2, vac_salary_years.values(), width, label=f'з/п {vac}')
        ax.set_xticks(x_nums, salary_years.keys(), rotation='vertical')
        ax.legend(fontsize=8)
        ax.tick_params(axis='both', labelsize=8)
        ax.grid(True, axis='y')

        # 2 - Количество вакансий по годам
        ax = fig.add_subplot(222)
        ax.set_title('Количество вакансий по годам')
        ax.bar(x_list1, vacs_years.values(), width, label='количество вакансий')
        ax.bar(x_list2, vac_count_years.values(), width, label=f'Количество вакансий\n{vac}')
        ax.set_xticks(x_nums, salary_years.keys(), rotation='vertical')
        ax.legend(fontsize=8)
        ax.tick_params(axis='both', labelsize=8)
        ax.grid(True, axis='y')

        # 3 - Уровень зарплат по городам
        width_y = 0.6
        y_nums = np.arange(len(res_sort_area_salary.keys()))
        y_list1 = y_nums

        ax = fig.add_subplot(223)
        ax.set_title('Уровень зарплат по городам')
        ax.barh(y_list1, Report.sym_n(res_sort_area_salary).values(), width_y, )
        ax.set_yticks(y_nums, Report.sym_n(res_sort_area_salary).keys())
        ax.tick_params(axis='x', labelsize=8)
        ax.tick_params(axis='y', labelsize=6)
        ax.grid(True, axis='x')
        plt.gca().invert_yaxis()

        # 4 - Доля вакансий по городам
        ax = fig.add_subplot(224)
        ax.set_title('Доля вакансий по городам')
        ax.pie(Report.plus_other(res_sort_fract_vac_area).values(),
               labels=Report.plus_other(res_sort_fract_vac_area).keys(), textprops={'fontsize': 6})

        plt.tight_layout()
        plt.savefig('graphLAST.png')
        plt.show()

    @staticmethod
    def weight(ws):
        """Изменяет ширину столбцов на автоматическую

        :param ws: Ecxel лист
        """
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

    @staticmethod
    def border(ws):
        """Создает границы ячеек

        :param ws: Ecxel лист
        """
        thin = Side(border_style='thin', color="000000")
        for i, column in enumerate(ws.columns):
            for cell in column:
                if cell.value != '':
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    @staticmethod
    def get_head(ws, columns):
        """Заполняет ячейки названием колонок

        :param ws: Ecxel лист
        :param columns:(list) Колонки
        """
        for i, column in enumerate(columns):
            ws.cell(row=1, column=(1 + i), value=column).font = Font(bold=True)

    @staticmethod
    def get_column_year(ws, info):
        """Заполняет колонки первого листа информацией с годами

        :param ws: Ecxel лист
        :param info: (tuple[dict]) статистика
        """
        salary_years = info[0]
        vac_salary_years = info[1]
        vacs_years = info[2]
        vac_count_years = info[3]
        for year, value in salary_years.items():
            ws.append([year, value, vac_salary_years[year], vacs_years[year], vac_count_years[year]])

    @staticmethod
    def get_column_area(ws, info):
        """Заполняет колонки второго листа информацией с городами

        :param ws: Ecxel лист
        :param info: (tuple[dict]) статистика
        """
        res_sort_area_salary = info[4]
        res_sort_fract_vac_area = info[5]
        # Добавление колонок города
        for i, area in enumerate(res_sort_area_salary.keys()):
            ws.cell(row=2 + i, column=1).value = area
            ws.cell(row=2 + i, column=3).value = ''
        for i, area in enumerate(res_sort_fract_vac_area.keys()):
            ws.cell(row=2 + i, column=4).value = area
        # Добавление колонки уровень зп
        for i, salary in enumerate(res_sort_area_salary.values()):
            ws.cell(row=2 + i, column=2).value = salary
        # Добавление колонки доля вакансий
        for i, fraction in enumerate(res_sort_fract_vac_area.values()):
            ws.cell(row=2 + i, column=5, value=fraction).number_format = BUILTIN_FORMATS[10]

    @staticmethod
    def report_excel(info, vac):
        """Генерирует файл Ecxel с таблицами по статистике

        :param info: (list[dict]): список статистики
        :param vac: (str): название профессии
        """
        columns1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {vac}', 'Количество вакансий',
                    f'Количество вакансий - {vac}']
        columns2 = ['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Статистика по годам'
        Report.get_head(ws1, columns1)
        Report.get_column_year(ws1, info)
        Report.weight(ws1)
        Report.border(ws1)

        ws2 = wb.create_sheet('Статистика по городам')
        Report.get_head(ws2, columns2)
        Report.get_column_area(ws2, info)
        Report.weight(ws2)
        Report.border(ws2)
        wb.save('report.xlsx')

    @staticmethod
    def generate_pdf(info, vac):
        """Генерирует файл pdf с графиками и таблицами по статистике

        :param info: (list[dict]): список статистики
        :param vac: (str): название профессии
        """
        columns1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {vac}', 'Количество вакансий',
                    f'Количество вакансий - {vac}']
        columns2 = ['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']
        salary_years = info[0]
        vac_salary_years = info[1]
        vacs_years = info[2]
        vac_count_years = info[3]
        res_sort_area_salary = info[4]
        res_sort_fract_vac_area = {}
        for key, value in info[5].items():
            res_sort_fract_vac_area[key] = str(round(info[5][key] * 100, 2)) + '%'
        del res_sort_fract_vac_area['Другие']
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("2.1.3.html")

        pdf_template = template.render({'vac': vac,
                                        'columns1': columns1,
                                        'salary_years': salary_years,
                                        'vac_salary_years': vac_salary_years,
                                        'vacs_years': vacs_years,
                                        'vac_count_years': vac_count_years,
                                        'res_sort_area_salary': res_sort_area_salary,
                                        'res_sort_fract_vac_area': res_sort_fract_vac_area,
                                        'columns2': columns2})

        options = {'enable-local-file-access': None}

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')

        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options=options)
